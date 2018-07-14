import openpyxl
import re
#Open "train.xlsx" file and open "train" sheet
open_workbook=openpyxl.load_workbook('./train.xlsx')
work_sheet=open_workbook['train']

#Make new excel file and create 4 sheets
new_workbook=openpyxl.Workbook()
new_worksheet_Mr=new_workbook.active
new_worksheet_Mr.title='남성'
for i in range(1,12+1):
    new_worksheet_Mr.cell(row=1,column=i).value=work_sheet.cell(row=1,column=i).value

new_worksheet_Miss=new_workbook.create_sheet('미혼여성')
for i in range(1,12+1):
    new_worksheet_Miss.cell(row=1,column=i).value=work_sheet.cell(row=1,column=i).value

new_worksheet_Mrs=new_workbook.create_sheet('기혼여성')
for i in range(1,12+1):
    new_worksheet_Mrs.cell(row=1,column=i).value=work_sheet.cell(row=1,column=i).value

new_worksheet_Others=new_workbook.create_sheet('기타')
for i in range(1,12+1):
    new_worksheet_Others.cell(row=1,column=i).value=work_sheet.cell(row=1,column=i).value

new_worksheet_Report=new_workbook.create_sheet('보고서')
new_worksheet_Report.append(['분류','생존자수','사망자수','생존률'])

Mrsurv=0
Mrnonsurv=0
Misssurv=0
Missnonsurv=0
Mrssurv=0
Mrsnonsurv=0
Othersurv=0
Othernonsurv=0

#Classify the gender data in "train.xlsx" and push each data(row) to the corresponding sheet in the new excel file
pattern=re.compile('[A-Za-z]+\.')
for each_row in work_sheet.rows:
    if len(pattern.findall(each_row[3].value))>0:
        if pattern.findall(each_row[3].value)[0]=='Mr.':
            new_worksheet_Mr.append([cell.value for cell in each_row])
            if each_row[1].value==1:
                Mrsurv+=1
            else:
                Mrnonsurv+=1
        elif pattern.findall(each_row[3].value)[0]=='Miss.':
            new_worksheet_Miss.append([cell.value for cell in each_row])
            if each_row[1].value==1:
                Misssurv+=1
            else:
                Missnonsurv+=1
        elif pattern.findall(each_row[3].value)[0]=='Mrs.':
            new_worksheet_Mrs.append([cell.value for cell in each_row])
            if each_row[1].value==1:
                Mrssurv+=1
            else:
                Mrsnonsurv+=1
        else:
            new_worksheet_Others.append([cell.value for cell in each_row])
            if each_row[1].value==1:
                Othersurv+=1
            else:
                Othernonsurv+=1

#Make a sheet that consists of the survival rate 
for i,j,k in [('남성',Mrsurv,Mrnonsurv),('미혼여성',Misssurv,Missnonsurv),('기혼여성',Mrssurv,Mrsnonsurv),('기타',Othersurv,Othernonsurv)]:
    new_worksheet_Report.append([i,j,k,"{0:.2f}%".format(100*(j/(j+k)))])

open_workbook.close()
new_workbook.save('./assginment_challenge.xlsx')
